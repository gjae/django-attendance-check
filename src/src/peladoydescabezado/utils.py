from decimal import Decimal
from typing import Optional, Union, List
from datetime import datetime, timedelta
from .models import TURNS_TIME_RANGES, Person, BasketProduction, TURNS, CATEGORIES

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from isoweek import Week

# Configurar el estilo de borde
border_style = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
font_style = Font(name='Arial Narrow', size=20, bold=True)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
font_style_subheader = Font(name='Arial Narrow', size=11, bold=True)
WEEKDAYS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

def get_current_turn(load_turn = None) -> Optional[str]:
    # Obtener la hora actual
    ahora = datetime.now().time()

    if load_turn is not None:
        if load_turn == 0:
            return "morning"
        return "night"

    for turn_key in TURNS_TIME_RANGES:
        start, end = TURNS_TIME_RANGES[turn_key]
        # Definir los límites del horario
        hora_inicio = datetime.strptime(start, "%H:%M").time()
        hora_fin = datetime.strptime(end, "%H:%M").time()
        
        # Verificar si la hora actual está dentro del rango
        if hora_inicio <= ahora <= hora_fin:
            return turn_key
        
    return "morning"


def _sheet_header(ws, turn, days, users, process):  
    part_flags = False
    current_day = 0
    ws['B1'] = f"RELACIÓN DIARIA DE PERSONAL A {process} ({TURNS[turn]})"
    # 1. Fusionar celdas desde A1 hasta BF4
    ws.merge_cells('A1:A4')
    ws.column_dimensions['A'].width = 30 # Ancho para la imagen
    img = None
    try:
        img = Image("/app/src/static/images/branding/logo_inpromaro_lit_backup.png")  # Cambia por tu archivo de imagen
        img.width = 100  # Ancho en píxeles
        img.height = 100  # Alto en píxeles
        ws.add_image(img, 'A1')
    except FileNotFoundError:
        print("Advertencia: No se encontró el archivo de imagen. Se insertará texto alternativo.")
        ws['A1'].value = "[LOGO]"
        ws['A1'].font = Font(size=20, bold=True, color='808080')
        ws['A1'].alignment = Alignment(vertical='center')

    ws["A1"].border =border_style
    ws.merge_cells('B1:K4')
    ws['B1'].font = font_style
    ws['B1'].alignment = center_alignment
    ws['B1'].border = border_style

    # Columnas iniciales
    start_col = 21  # BI (columna 61)
    end_col = len(days) + 100   # EK (columna 135)
    block_size = 8  # BI a BP (8 columnas: BI,BJ,BK,BL,BM,BN,BO,BP)
    skip_size = 7  # Columnas a ignorar (BQ,BR,BS,BT)
    
    current_col = start_col
    
    while current_col <= end_col and part_flags:
        # 2. Ignorar BH (columna 60) - ya está fuera del rango inicial
        
        # 3. Fusionar BI a BP (8 columnas)
        if current_col + block_size - 1 <= end_col:
            if current_day >= len(days):
                break

            current_date = days[current_day]
            start_letter = get_column_letter(current_col)
            end_letter = get_column_letter(current_col + block_size - 1)
            merge_range = f"{start_letter}1:{end_letter}4"
            
            ws.merge_cells(merge_range)
            merged_cell = ws[f"{start_letter}1"]
            merged_cell.value = f"RELACIÓN DIARIA DE PERSONAL A DESTAJO ({process})"
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            merged_cell.border = border_style
            merged_cell.font = font_style_subheader
            if img is not None:
                i = Image("/app/src/static/images/branding/logo_inpromaro_lit_backup.png")
                        
                i.width = 100  # Ancho en píxeles
                i.height = 80  # Alto en píxeles
                ws.add_image(i, f"{start_letter}1")

            # Apartado de dia y fecha
            ws[f"{get_column_letter(current_col + block_size)}1"] = WEEKDAYS[current_date.weekday()]
            ws[f"{get_column_letter(current_col + block_size)}1"].alignment = center_alignment
            ws[f"{get_column_letter(current_col + block_size)}1"].font = font_style_subheader
            ws.merge_cells(f"{get_column_letter(current_col + block_size)}1:{get_column_letter(current_col + block_size + 2)}2")

            #sub apartado de fecha
            ws[f"{get_column_letter(current_col + block_size)}3"] = current_date.day
            ws[f"{get_column_letter(current_col + block_size + 1)}3"] = current_date.month
            ws[f"{get_column_letter(current_col + block_size + 2)}3"] = current_date.year
            ws[f"{get_column_letter(current_col + block_size)}3"].alignment = center_alignment
            ws[f"{get_column_letter(current_col + block_size)}3"].font = font_style_subheader
            ws[f"{get_column_letter(current_col + block_size + 1)}3"].alignment = center_alignment
            ws[f"{get_column_letter(current_col + block_size + 1)}3"].font = font_style_subheader
            ws[f"{get_column_letter(current_col + block_size + 2)}3"].alignment = center_alignment
            ws[f"{get_column_letter(current_col + block_size + 2)}3"].font = font_style_subheader
            

            ws.merge_cells(f"{get_column_letter(current_col + block_size)}3:{get_column_letter(current_col + block_size)}4")
            ws.merge_cells(f"{get_column_letter(current_col + block_size + 1)}3:{get_column_letter(current_col + block_size + 1)}4")
            ws.merge_cells(f"{get_column_letter(current_col + block_size + 2)}3:{get_column_letter(current_col + block_size + 2)}4")
        
            current_day += 1
        # Avanzar el contador (8 columnas fusionadas + 4 ignoradas)
        current_col += block_size + skip_size
    
    # Ajustar anchos de columna
    if part_flags:
        for col in range(1, 136):  # Desde A hasta EK
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        ws.column_dimensions["A"].width = 34

    for (letter, content) in [("A", "APELLIDO Y NOMBRE"), ("B", "CÉDULA"), ("C", "TURNO")]:
        ws[f"{letter}6"] = content
        ws[f"{letter}6"].font = font_style_subheader
        ws[f"{letter}6"].alignment = center_alignment
        ws.merge_cells(f"{letter}6:{letter}7")


    sheet_name_index = 8
    ids = set()
    for u in users:
        if u['turn'] != turn:
            continue
        if u['identity'] in ids:
            continue

        ids.add(u['identity'])
        ws[f"A{sheet_name_index}"].value = f"{u['lastnames']}, {u['names']}"
        ws[f"B{sheet_name_index}"].value = f"{u['identity']}"
        ws[f"C{sheet_name_index}"].value = f"{TURNS[u['turn']]}"

        ws[f"A{sheet_name_index}"].alignment = center_alignment
        ws[f"B{sheet_name_index}"].alignment = center_alignment
        ws[f"C{sheet_name_index}"].alignment = center_alignment 
        sheet_name_index += 1

def _apply_font_cell_format(cell, font, alignment, value, width = None, letter = None, sheet = None, border = None):
    cell.value = value
    cell.font = font
    cell.alignment = alignment
    if width is not None and letter is not None and sheet is not None:
        sheet.column_dimensions[letter].width = width

    if border is not None and isinstance(border, Border):
        cell.border = border

def get_days_between_dates(start_date: Union[str, datetime], 
                          end_date: Union[str, datetime],
                          format_output: str = None) -> List[datetime]:
    """
    Retorna una lista de días entre dos fechas (inclusive).
    
    Args:
        start_date: Fecha inicial (str en formato 'YYYY-MM-DD' o objeto datetime)
        end_date: Fecha final (str en formato 'YYYY-MM-DD' o objeto datetime)
        format_output: Si se especifica (ej: '%Y-%m-%d'), retorna strings formateados
    
    Returns:
        Lista de objetos datetime (o strings si format_output está definido)
    
    Raises:
        ValueError: Si las fechas no son válidas o start_date > end_date
    """
    # Convertir strings a datetime si es necesario
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Asegurarse de que son objetos date (sin hora)
    if hasattr(start_date, 'date'):
        start_date = start_date.date()
    if hasattr(end_date, 'date'):
        end_date = end_date.date()
    
    # Validar rango
    if start_date > end_date:
        raise ValueError("La fecha de inicio no puede ser posterior a la fecha final")
    
    # Generar rango de días
    days = []
    current_date = start_date
    while current_date <= end_date:
        if format_output:
            days.append(current_date.strftime(format_output))
        else:
            days.append(current_date)
        current_date += timedelta(days=1)
    
    return days

def _get_xlsx_report_template(date_from = None, date_end = None, process = "DESCABEZADO", turn = 0, wb = None, ws = None, category = 3):
    # Crear un nuevo libro de trabajo
    part_flags = False
    days = get_days_between_dates(date_from, date_end)
    production, users, daily_total = BasketProduction.objects.weekly_record(days[0], days[-1], category=category)
    daily_block = {}

    start_letter_pick = 21
    for i, day in enumerate(days):
        letter = get_column_letter(start_letter_pick)
        end_letter = get_column_letter(start_letter_pick + 10)
        daily_block[day] = {"letter_start": (letter, start_letter_pick), "letter_end": (end_letter, start_letter_pick + 10)}
        start_letter_pick += 15

 
    
    _sheet_header(ws, turn, days, users, process)
    if part_flags:
        for day in days:
            block = daily_block[day]
            cell_letter, cell_name = block["letter_start"]
            _apply_font_cell_format(
                ws[f"{get_column_letter(cell_name)}6"], 
                font_style_subheader, 
                center_alignment, 
                "APELLIDOS Y NOMBRES", 
                width=30, 
                letter=cell_letter, 
                sheet=ws
            )
            _apply_font_cell_format(ws[f"{get_column_letter(cell_name + 1)}6"], font_style_subheader, center_alignment, "PROCESO")
            _apply_font_cell_format(ws[f"{get_column_letter(cell_name + 2)}6"], font_style_subheader, center_alignment, "TURNO")
            _apply_font_cell_format(ws[f"{get_column_letter(cell_name + 3)}6"], font_style_subheader, center_alignment, "CÉDULA")
            _apply_font_cell_format(ws[f"{get_column_letter(cell_name + 4)}6"], font_style_subheader, center_alignment, "FECHA")
            _apply_font_cell_format(ws[f"{get_column_letter(cell_name + 5)}6"], font_style_subheader, center_alignment, "TOTAL PESADO" if category != 3 else "TOTAL CESTAS")


    last_cell = None
    for i, day in enumerate(days):
        if last_cell is None:
            last_cell = i + 4
        start_letter = get_column_letter(last_cell)
        last_cell += 1
        end_letter = get_column_letter(last_cell)
        _apply_font_cell_format(
            ws[f"{start_letter}6"],
            font_style_subheader,
            center_alignment,
            f"{WEEKDAYS[day.weekday()]} ({day.strftime('%d/%m/%Y')})",
            width=20,
            sheet=ws,
            letter=start_letter
        )

        ws[f"{start_letter}7"] = "Total" if category != 3 else "Total (cestas)"
        ws[f"{start_letter}7"].font = font_style_subheader
        ws[f"{start_letter}7"].alignment = center_alignment
        ws.column_dimensions["B"].width = 20
        
        start_totalization_cell = 4
        skip_cells = 1
        total_personal = {}
        cell = None
        for day in days:
            cell = get_column_letter(start_totalization_cell)

            current_totalization_row = 8
            ids = set()
            weight_totals = 0
            basket_totals = 0
            for u in users:
                if u['identity'] in ids:
                    continue

                ids.add(u['identity'])
                if str(u['identity']) not in production[turn][day]:
                    ws[f"{cell}{current_totalization_row}"] = Decimal(0.00)
                    ws[f"{cell}{current_totalization_row}"].alignment = center_alignment
                    current_totalization_row += 1
                    continue
                
                weight_totals += production[turn][day][str(u['identity'])]['weight_sum']
                basket_totals += production[turn][day][str(u['identity'])]['basket_count']
                ws[f"{cell}{current_totalization_row}"] = f"{production[turn][day][str(u['identity'])]['weight_sum']}" if category != 3 else production[turn][day][str(u['identity'])]['basket_count']
                ws[f"{cell}{current_totalization_row}"].alignment = center_alignment
                identity = str(u['identity'])
                if identity not in total_personal:
                    total_personal[identity] = (current_totalization_row, 0)

                total_personal[identity] = (current_totalization_row, total_personal[identity][1] + (production[turn][day][identity]['weight_sum'] if category != 3 else production[turn][day][str(u['identity'])]['basket_count']))
                current_totalization_row += 1

            start_totalization_cell += skip_cells
            ws[f"{cell}{current_totalization_row+1}"] = weight_totals if category != 3 else basket_totals
            ws[f"{cell}{current_totalization_row+1}"].alignment = center_alignment

        ws[f"A{current_totalization_row+1}"] = "Total por día" if category != 3 else "Total por día de cestas"
        ws[f"A{current_totalization_row+1}"].font = Font(bold=True, size=14)
        ws[f"A{current_totalization_row+1}"].alignment = Alignment(vertical="center", horizontal="right")
        ws.merge_cells(f"A{current_totalization_row+1}:C{current_totalization_row+1}")

        _apply_font_cell_format(
            ws[f"{get_column_letter(start_totalization_cell)}6"],
            font_style_subheader,
            center_alignment,
            "Total por trabajador",
            width=24,
            letter=get_column_letter(start_totalization_cell ),
            sheet=ws
        )
        if cell is not None:
            cell = get_column_letter(start_totalization_cell)
            for row, total in total_personal.values():
                ws[f"{cell}{row}"] = total
            

        
            

def get_xlsx_report_template(date_from = None, date_end = None, turn = 0):
    wb = Workbook()
    sheet = None
    sheet = wb.active
    categories = dict(CATEGORIES)
    
    i = 0
    for cat_key in categories:
        process = categories[cat_key]
        if i == 0:
            aux = f"Auxiliar {process} ({TURNS[turn]})"
            sheet.title = aux
        else:
            sheet = wb.create_sheet(f"Auxiliar {process} ({TURNS[turn]})")
        _get_xlsx_report_template(date_from = date_from, date_end = date_end, process = process, turn=turn, wb=wb, ws=sheet, category=cat_key)
        i += 1

    return wb


def generate_rport_xlsx_simple(context):
    wb = Workbook()
    ws = wb.active
    # /Header

    _apply_font_cell_format(ws["B1"], font_style_subheader, center_alignment, f"CONTROL DE ENTREGA VALOR AGREGADO ({context['category_text']})",)
    _apply_font_cell_format(ws["G1"], font_style_subheader, center_alignment, f"Código: {context['code']}")
    _apply_font_cell_format(ws["G2"], font_style_subheader, center_alignment, f"Fecha: {context['date'].strftime('%d/%m/%Y')}")
    _apply_font_cell_format(ws["G3"], font_style_subheader, center_alignment, f"Versión: {context['version']}")

    _apply_font_cell_format(ws["A14"], font_style_subheader, center_alignment, f"N°")
    _apply_font_cell_format(ws["B14"], font_style_subheader, center_alignment, f"C.I")
    _apply_font_cell_format(ws["C14"], font_style_subheader, center_alignment, f"APELLIDO Y NOMBRE")
    _apply_font_cell_format(ws["D14"], font_style_subheader, center_alignment, f"Contenido de cestas entregadas")
    
    ws["A6"] = f"Fecha: {context['date'].strftime('%d/%m/%Y')}"
    ws["B6"] = f"Presentacion: "
    ws["C6"] = "Cola: "
    ws["D6"] = "Cola Pelada: "
    ws["E6"] = "Especie: "

    content_letter_end = get_column_letter(context["num_max_totalization_cells"] + 3)
    content_total_col = get_column_letter(context['num_max_totalization_cells'] + 4)
    _apply_font_cell_format(ws[f"{content_total_col}14"], font_style_subheader, center_alignment, f"T. Cestas" if context['category'] == 3 else "T. Kgs")
    
    col_detail_start = 1
    for control in context["controls"]:
        for detail in control.control_details:
            col_letter = get_column_letter(col_detail_start)
            ws[f"{col_letter}7"] = "Producto certificado: "
            ws[f"{col_letter}8"] = f"Granja: {detail.farm.name}"
            ws[f"{col_letter}9"] = f"Piscina: {detail.pool.number}"
            ws[f"{col_letter}10"] = f"Gramaje:  {', '.join([w.weight for w in detail.weightness.all()])}"
            ws[f"{col_letter}11"] = f"Kilos de camaron entero recibidos: {detail.total_weight_received}"
            col_detail_start += 1
    
    
    # 1. Fusionar celdas desde A1 hasta BF4
    ws.merge_cells('A1:A4')
    ws.column_dimensions['A'].width = 20 # Ancho para la imagen
    img = None
    try:
        img = Image("/app/src/static/images/branding/logo_inpromaro_lit_backup.png")  # Cambia por tu archivo de imagen
        img.width = 100  # Ancho en píxeles
        img.height = 100  # Alto en píxeles
        ws.add_image(img, 'A1')
    except FileNotFoundError:
        print("Advertencia: No se encontró el archivo de imagen. Se insertará texto alternativo.")
        ws['A1'].value = "[LOGO]"
        ws['A1'].font = Font(size=20, bold=True, color='808080')
        ws['A1'].alignment = Alignment(vertical='center')

    ws.merge_cells("B1:F4")
    ws.merge_cells("G1:K1")
    ws.merge_cells("G2:K2")
    ws.merge_cells("G3:K4")
    if context["num_max_totalization_cells"] > 0:
        ws.merge_cells(f"D14:{content_letter_end}14")

    if context["num_max_totalization_cells"] < 3:
        ws.column_dimensions["D"].width = 30

    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions[f"{content_total_col}"].width = 20
    # Body
    row = 15
    for user in context["data"]:
        ws[f"A{row}"] = user.row
        ws[f"B{row}"] = user.identity
        ws[f"C{row}"] = f"{user.lastnames}, {user.names}"
        ws[f"{content_total_col}{row}"] = user.total if user.total else 0
        col = 4
        for counter in context["num_cells"]:
            if counter < len(user.production):
                ws.cell(row=row, column=col, value=user.production[counter].weight)
            else:
                break
            col += 1
        row += 1
    ws[f"{content_total_col}{row}"] =  "Total Kg" if context["category"] != 3 else "Total cestas" 
    ws[f"{content_total_col}{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"{content_total_col}{row}"].font = Font(bold=True, size=10, name="Arial Narrow")
    row += 1
    ws[f"{content_total_col}{row}"] = context["weight_total"]

    ws[f"A{row+1}"].value = "Elaborado por: "
    ws[f"B{row+1}"].value = context.get("created_by_names", "")
    # output
    return wb


def attendance_by_department_xlsx(context):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Reporte por departamento"

    # header
    try:
        img = Image("/app/src/static/images/branding/logo_inpromaro_lit_backup.png")  # Cambia por tu archivo de imagen
        img.width = 100  # Ancho en píxeles
        img.height = 100  # Alto en píxeles
        sheet.add_image(img, 'A1')
    except FileNotFoundError:
        print("Advertencia: No se encontró el archivo de imagen. Se insertará texto alternativo.")
        sheet['A1'].value = "[LOGO]"
        sheet['A1'].font = Font(size=20, bold=True, color='808080')
        sheet['A1'].alignment = Alignment(vertical='center')

    sheet.column_dimensions["A"].width = 20
    sheet.merge_cells("A1:A6")
    _apply_font_cell_format(
        sheet["B1"],
        font_style_subheader,
        center_alignment,
        f"Reporte por departamento",
    )
    sheet.merge_cells("B1:H1")
    _apply_font_cell_format(
        sheet["B2"],
        font_style_subheader,
        center_alignment,
        f"Departamento: {context['department'].name}"
    )
    sheet.merge_cells("B2:H2")
    _apply_font_cell_format(
        sheet["B3"],
        font_style_subheader,
        center_alignment,
        f"Desde {context['range_start'].strftime('%d/%m/%Y')} hasta {context['range_end'].strftime('%d/%m/%Y')}"
    )
    sheet.merge_cells("B3:H3")
    _apply_font_cell_format(
        sheet["B4"],
        font_style_subheader,
        center_alignment,
        f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y')}"
    )
    sheet.merge_cells("B4:H6")

    # header sheet 2
    sheet_obs = wb.create_sheet("Observaciones")
    o_cols = [
        "Cédula", 
        "Nombre",
        "Apellido",
        "Fecha",
        "Descripción"
    ]
    for index, col in enumerate(o_cols):
        letter = get_column_letter(index+1)
        sheet_obs[f"{letter}1"] = col
        sheet_obs.column_dimensions[f"{letter}"].width = 35
        sheet_obs[f"{letter}1"].border = border_style



    # body
    #body sheet 1
    cols = [
        "Cédula", 
        "Nombre",
        "Apellido",
        "Cargo",
        "Días Marcados",
        "Total de horas"
    ]

    for index, col in enumerate(cols):
        _apply_font_cell_format(
            sheet[f"{get_column_letter(index+1)}8"],
            font_style_subheader,
            center_alignment,
            col,
            width=40,
            sheet=sheet,
            letter=get_column_letter(index+1),
            border=border_style
        )

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["E"].width = 20

    row = 9
    total_hours = 0
    for user in context['data']:
        sheet[f"A{row}"] = user["person"].identity
        sheet[f"B{row}"] = user["person"].names
        sheet[f"C{row}"] = user["person"].lastnames
        sheet[f"D{row}"] = user["person"].position.position
        sheet[f"E{row}"] = context["total_days_by_user"].get(user["person"].id, "--")
        sheet[f"F{row}"] = user["abs_total_hours"]
        total_hours += user["abs_total_hours"]
        row += 1
    
    sheet[f"A{row}"] = "Horas trabajadas"
    sheet[f"F{row}"] = total_hours
    for index, col in enumerate(cols):
        sheet[f"{get_column_letter(index+1)}{row}"].border = border_style


    # body observations
    row = 2
    for observation in context["observations"]:
        sheet_obs[f"A{row}"] = observation.person.identity
        sheet_obs[f"B{row}"] = observation.person.names
        sheet_obs[f"C{row}"] = observation.person.lastnames
        sheet_obs[f"D{row}"] = observation.calendar_day.date_day.strftime("%d/%m/%Y")
        sheet_obs[f"E{row}"] = observation.description
        row += 1

    return wb



def attendance_by_personal_xlsx(context):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Reporte por departamento"

    # header
    try:
        img = Image("/app/src/static/images/branding/logo_inpromaro_lit_backup.png")  # Cambia por tu archivo de imagen
        img.width = 100  # Ancho en píxeles
        img.height = 100  # Alto en píxeles
        sheet.add_image(img, 'A1')
    except FileNotFoundError:
        print("Advertencia: No se encontró el archivo de imagen. Se insertará texto alternativo.")
        sheet['A1'].value = "[LOGO]"
        sheet['A1'].font = Font(size=20, bold=True, color='808080')
        sheet['A1'].alignment = Alignment(vertical='center')

    sheet.column_dimensions["A"].width = 20
    sheet.merge_cells("A1:A6")
    _apply_font_cell_format(
        sheet["B1"],
        font_style_subheader,
        center_alignment,
        f"Reporte por persona",
    )
    sheet.merge_cells("B1:H1")
    _apply_font_cell_format(
        sheet["B2"],
        font_style_subheader,
        center_alignment,
        f"Departamento: {context['person'].department.name}"
    )
    sheet.merge_cells("B2:H2")
    _apply_font_cell_format(
        sheet["B3"],
        font_style_subheader,
        center_alignment,
        f"Trabajador {context['person'].get_fullname()} - {context['person'].identity}"
    )
    sheet.merge_cells("B3:H3")
    _apply_font_cell_format(
        sheet["B4"],
        font_style_subheader,
        center_alignment,
        f"Desde {context['range_start'].strftime('%d/%m/%Y')} hasta {context['range_end'].strftime('%d/%m/%Y')}"
    )
    sheet.merge_cells("B4:H4")
    _apply_font_cell_format(
        sheet["B5"],
        font_style_subheader,
        center_alignment,
        f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y')}"
    )
    sheet.merge_cells("B5:H6")

    # header sheet 2
    sheet_obs = wb.create_sheet("Observaciones")
    o_cols = [
        "Cédula", 
        "Nombre",
        "Apellido",
        "Fecha",
        "Descripción"
    ]
    for index, col in enumerate(o_cols):
        letter = get_column_letter(index+1)
        sheet_obs[f"{letter}1"] = col
        sheet_obs.column_dimensions[f"{letter}"].width = 35
        sheet_obs[f"{letter}1"].border = border_style



    # body
    #body sheet 1
    cols = [
        "Cédula", 
        "Nombre",
        "Apellido",
        "Observaciones",
        "Total de horas"
    ]

    for index, col in enumerate(cols):
        _apply_font_cell_format(
            sheet[f"{get_column_letter(index+1)}8"],
            font_style_subheader,
            center_alignment,
            col,
            width=40,
            sheet=sheet,
            letter=get_column_letter(index+1),
            border=border_style
        )

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["E"].width = 20

    row = 9
    total_hours = 0
    for user in context['data']:
        sheet[f"A{row}"] = user["person"].identity
        sheet[f"B{row}"] = user["person"].names
        sheet[f"C{row}"] = user["person"].lastnames
        sheet[f"D{row}"] = "Si" if len(context["observations"]) > 0 else "No"
        sheet[f"E{row}"] = user["abs_total_hours"]
        total_hours += user["abs_total_hours"]
        row += 1
    
    sheet[f"A{row}"] = "Horas trabajadas"
    sheet[f"E{row}"] = total_hours
    for index, col in enumerate(cols):
        sheet[f"{get_column_letter(index+1)}{row}"].border = border_style


    # body observations
    row = 2
    for observation in context["observations"]:
        sheet_obs[f"A{row}"] = observation.person.identity
        sheet_obs[f"B{row}"] = observation.person.names
        sheet_obs[f"C{row}"] = observation.person.lastnames
        sheet_obs[f"D{row}"] = observation.calendar_day.date_day.strftime("%d/%m/%Y")
        sheet_obs[f"E{row}"] = observation.description
        row += 1

    return wb
