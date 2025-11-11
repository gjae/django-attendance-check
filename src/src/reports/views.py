import calendar
from datetime import datetime
from django.conf import settings
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from django.views.generic.base import ContextMixin
from django.http.response import HttpResponse
from django_weasyprint import WeasyTemplateResponseMixin
from django.views.generic import TemplateView, DetailView, View
from django.db.models import Window, Count, Q, Prefetch
from django.db.models.functions import RowNumber
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font


from src.employees.models import Employee
from src.reports.models import TimeReport
from src.reports.utils import has_observation
from src.settings.models import Department
from src.clocking.models import DailyChecks, DailyCalendarObservation
from src.settings.models import WorkCenter


class ReportLetterheadException(Exception):
    pass


class ReportBrandingLogoException(Exception):
    pass

class BaseReportMixin(ContextMixin):
    """
    Clase mixin base para la generación de reportes
    debe especificarse una propiedad letterhead_lines que será el
    encargado mostrar el membrete
    """

    def get_letterhead_lines(self):
        if not hasattr(self, "letterhead_lines"):
            raise ReportLetterheadException("letterhead_lines doesn't define")
        
        if not isinstance(self.letterhead_lines, list) and not isinstance(self.letterhead_lines, tuple):
            raise ReportLetterheadException("letterhead_lines maybe iterable object")
        

        return self.letterhead_lines
    
    def get_branding_logo(self):
        if not hasattr(self, "branding_logo") or self.branding_logo == "":
            raise ReportBrandingLogoException("branding_logo must be defined")
        
        return self.branding_logo
    
    def get_report_type(self):
        return self.report_type


    def get_context_data(self):
        context = {
            "letterheads": self.get_letterhead_lines(),
            "branding": self.get_branding_logo(),
            "report_type": self.get_report_type()
        }

        return context

class ReportByWorkerPdfView(BaseReportMixin, WeasyTemplateResponseMixin, TemplateView):
    template_name = "reports/by_worker.pdf.html"
    pdf_attachment = False
    report_type = "trabajador"
    branding_logo = "/app/src/static/images/branding/logo_inpromaro_lit.png"
    pdf_stylesheets = [
        '/app/src/static/css/bootstrap.min.css',
    ]
    letterhead_lines = (
        "INPROMAR C.A",
        "Reporte de asistencia por trabajador",
    )
    extra_context = {"show_metadata": True}
    employer = None


    def get_pdf_filename(self):
        now = timezone.now()
        return (
            "trabajador-{at}.pdf".format(at=now.strftime("%d-%m-%Y"), )
        )
    

    def get_letterhead_lines(self):
        now = timezone.now()
        letterhead = super().get_letterhead_lines()
        self.employer = Employee.objects.select_related().get(id=int(self.request.GET.get("employer")))

        return letterhead + (
            "<strrong>Fecha de generación {date}</strong>".format(date=now.strftime("%d/%m/%Y %I:%M %p")),
        ) + (
            "Reporte del trabajador: {}".format(self.employer.get_fullname()),
            "Desde <strong>{}</strong> hasta <strong>{}</strong>".format(
                datetime.strptime(self.request.GET.get("start_at"), "%Y-%m-%d").strftime("%d/%m/%Y"),
                datetime.strptime(self.request.GET.get("end_at"), "%Y-%m-%d").strftime("%d/%m/%Y"),
            )
        )
    
    def get_context_data(self):
        context = super().get_context_data()
        data, total_hours, _ = DailyChecks.objects.report_by_employee(
            int(self.request.GET.get("employer")), 
            self.request.GET.get("start_at"), 
            self.request.GET.get("end_at")
        )

        context["data"] = data
        context["total_hours"] = total_hours
        context["employer"] = Employee.objects.select_related().get(id=int(self.request.GET.get("employer"))) if self.employer is None else self.employer
        context["range_start"] = datetime.strptime(self.request.GET.get("start_at"), "%Y-%m-%d")
        context["range_end"] = datetime.strptime(self.request.GET.get("end_at"), "%Y-%m-%d")
        context["observations"] = DailyCalendarObservation.objects.select_related("employer", "calendar_day").filter(
            employer_id=int(self.request.GET.get("employer")),
            calendar_day__date_day__range=[ 
                self.request.GET.get("start_at"), 
                self.request.GET.get("end_at")
            ]
        )

        # context de inasistencias del empleado
        context["absences"] = Employee.objects.absences_between_dates(self.request.GET.get("start_at"), self.request.GET.get("end_at"))

        return context
    
    def post(self, request, *args, **kwargs):
        return self.render_to_response(self.get_context_data())


class ReportByWorkerView(LoginRequiredMixin, TemplateView):
    template_name = "reports/by_worker.html"


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        workcenters = []
        workcenter_objects = WorkCenter.objects.prefetch_related(
            Prefetch(
                "departments",
                queryset=Department.objects.select_related("work_center").filter(is_actived=True).prefetch_related("employers").order_by("name"),
                to_attr="all_departments"
            )
        )
        for w in workcenter_objects:
            workcenters.append({
                "id": w.id,
                "name": w.name,
                "departments": [{
                    "id": d.id,
                    "name": str(d),
                    "users": [{
                        "id": u.id,
                        "name": u.get_fullname(),
                        "cedula": u.cedula,
                    } for u in d.employers.all()]
                } for d in w.all_departments]
            })


        context["employers"] = Employee.objects.only_actives().order_by("last_name", "name")
        context["departments"] = workcenters
        context["annuities"] = [2024, 2025, 2026]
        context["current_annuity"] = 2025
        context["current_month"] = datetime.now().month
        return context


class ReportByDepartmentPdfView(BaseReportMixin, WeasyTemplateResponseMixin, TemplateView):
    template_name = "reports/by_department.pdf.html"
    pdf_attachment = False
    report_type = "trabajador"
    branding_logo = "/app/src/static/images/branding/logo_inpromaro_lit.png"
    pdf_stylesheets = [
        '/app/src/static/css/bootstrap.min.css',
    ]
    letterhead_lines = (
        "INPROMAR C.A",
        "Reporte de asistencia por departamento",
    )
    department = None


    def get_pdf_filename(self):
        now = timezone.now()
        return (
            "departamento-{at}.pdf".format(at=now.strftime("%d-%m-%Y"), )
        )
    

    def get_letterhead_lines(self):
        now = timezone.now()
        letterhead = super().get_letterhead_lines()
        self.department = Department.objects.select_related().get(id=int(self.request.GET.get("department")))
        
        return letterhead + (
            "<strrong>Fecha de generación {date}</strong>".format(date=now.strftime("%d/%m/%Y %I:%M %p")),
            "Departamento: {}".format(self.department.name),
            "Desde <strong>{}</strong> hasta <strong>{}</strong>".format(
                datetime.strptime(self.request.GET.get("start_at"), "%Y-%m-%d").strftime("%d/%m/%Y"),
                datetime.strptime(self.request.GET.get("end_at"), "%Y-%m-%d").strftime("%d/%m/%Y"),
            )
        )
    
    def get_context_data(self):
        context = super().get_context_data()
        data, total_hours, total_days_by_user = DailyChecks.objects.report_by_department(from_date=self.request.GET.get("start_at"), until_date=self.request.GET.get("end_at"), department=int(self.request.GET.get("department", "1")))

        context["data"] = data
        context["total_hours"] = total_hours
        context["department"] = Department.objects.select_related().get(id=int(self.request.GET.get("department"))) if self.department is None else self.department
        context["range_start"] = datetime.strptime(self.request.GET.get("start_at"), "%Y-%m-%d")
        context["range_end"] = datetime.strptime(self.request.GET.get("end_at"), "%Y-%m-%d")
        context["total_days_by_user"] = total_days_by_user
        context["observations"] = DailyCalendarObservation.objects.select_related("employer", "calendar_day").filter(
            employer__department_id=int(self.request.GET.get("department")),
            calendar_day__date_day__range=[ 
                self.request.GET.get("start_at"), 
                self.request.GET.get("end_at")
            ]
        )
        context["show_metadata"] = False
        return context
    
    def post(self, request, *args, **kwargs):
        return self.render_to_response(self.get_context_data())


class ReportByDepartmentView(LoginRequiredMixin, TemplateView):
    template_name = "reports/by_department.html"


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        workcenters = []
        workcenter_objects = WorkCenter.objects.prefetch_related(
            Prefetch(
                "departments",
                queryset=Department.objects.select_related("work_center").filter(is_actived=True).order_by("name"),
                to_attr="all_departments"
            )
        )
        for w in workcenter_objects:
            workcenters.append({
                "id": w.id,
                "name": w.name,
                "departments": [{
                    "id": d.id,
                    "name": str(d)
                } for d in w.all_departments]
            })


        context["departments"] = workcenters
        context["departments"] = workcenters
        context["annuities"] = [2024, 2025, 2026]
        context["current_annuity"] = 2025
        context["current_month"] = datetime.now().month
        return context


class AttendanceReport(LoginRequiredMixin, TemplateView):
    template_name = "reports/by_attendance.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        workcenters = []
        workcenter_objects = WorkCenter.objects.prefetch_related(
            Prefetch(
                "departments",
                queryset=Department.objects.select_related("work_center").filter(is_actived=True).order_by("name"),
                to_attr="all_departments"
            )
        )
        for w in workcenter_objects:
            workcenters.append({
                "id": w.id,
                "name": w.name,
                "departments": [{
                    "id": d.id,
                    "name": str(d)
                } for d in w.all_departments]
            })


        context["departments"] = workcenters
        context["annuities"] = [2024, 2025, 2026]
        context["current_annuity"] = 2025
        context["current_month"] = datetime.now().month
        return context

class ReportByAttendancePdfView(BaseReportMixin, WeasyTemplateResponseMixin, TemplateView):
    pdf_attachment = False
    report_type = "trabajador"
    branding_logo = "/app/src/static/images/branding/logo_inpromaro_lit.png"
    pdf_stylesheets = [
        '/app/src/static/css/bootstrap.min.css',
    ]
    letterhead_lines = [
        "INPROMAR C.A",
        "Reporte de asistencia por departamento",
    ]

    def get_template_names(self):
        if self.request.GET.get("type_report", "simple") == "simple":
            return ["reports/by_attendance.pdf.html", ]
        
        return ["reports/by_detail_attendance.pdf.html", ]

    def get_pdf_filename(self):
        now = timezone.now()
        kind = "departamento"
        if int(self.request.GET.get("work_center", 1)) != 1:
            kind = "empresa"

        if self.request.GET.get("type_report", "simple") == "detallado":
            kind = f"detallado_{kind}"

        return (
            f"{kind}-{now.strftime('%d-%m-%Y')}.pdf"
        )
    

    def get_letterhead_lines(self):
        now = timezone.now()
        letterhead = super().get_letterhead_lines()

        kind = ""
        
        if int(self.request.GET.get("work_center", 1)) != 1:
            kind = "por empresa"
            letterhead[1] = f"Reporte de asistencia {kind}"
        else:
            kind = "por departamento" if int(self.request.GET.get("by_office", 0)) == 1 else ""
            letterhead[1] = f"Reporte de asistencia {kind}"

        return letterhead + [
            "<strrong>Fecha de generación {date}</strong>".format(date=now.strftime("%d/%m/%Y %I:%M %p")),
        ]
    
    def get_records(self):
        from src.clocking.reportering import AttendanceReport
        department = None
        if int(self.request.GET.get("by_office", 0)) == 1:
            department = int(self.request.GET.get("department"))
        report = AttendanceReport(
            self.request.GET.get("start_at"), 
            self.request.GET.get("end_at"), 
            department=department, 
            include_unattendances=False,
            work_center=int(self.request.GET.get("work_center", 1))
        )

        if self.request.GET.get("type_report", "simple") == "simple":
            print("Records simple")
            return report.sort().unique_by_user().make(simple=True)[0]
        
        return dict(report.sort().group_by_date())
    
    def _get_simple_report(self):
        records = []
        context = super().get_context_data()
        records = self.get_records()

        if int(self.request.GET.get("by_office", 0)) == 1:
            context["department"] = Department.objects.filter(id=int(self.request.GET.get("department"))).first()

        context["data"] = records
        context["range_start"] = datetime.strptime(self.request.GET.get("start_at"), "%Y-%m-%d").strftime("%d/%m/%Y")
        context["range_end"] = datetime.strptime(self.request.GET.get("end_at"), "%Y-%m-%d").strftime("%d/%m/%Y")
        context["only_department"] = int(self.request.GET.get("by_office", 0)) == 1
        return context
    
    def get_context_data(self):
        print(f"Tipo de reporte ", self.request.GET)
        return self._get_simple_report()
    
    def post(self, request, *args, **kwargs):
        return self.render_to_response(self.get_context_data())


class ReportExcelMixin(LoginRequiredMixin, View):

    def get_footer(self, total_hours):
        num_fields = len(self.get_headers())
        fields = ["Horas trabajadas"]

        for i in range(num_fields - 2):
            fields.append(" ")

        fields.append(total_hours)

        return fields
    
    def before_process(self):
        pass

    def before(self, ws):
        pass

    def post_processing(self, data, sheet, wb):
        pass

    def set_header(self, ws):
        headers = self.get_headers()
        ws.append(headers)

    def set_sheet_title(self, ws, title = None):
        if title is None:
            ws.title = self.get_sheet_title()
        else:
            ws.title = title
    
    def get_worksheet(self):
        workbook = Workbook()
        ws = workbook.active
        return workbook, ws
    
    def fit_column_size(self, ws):
        dim_holder = DimensionHolder(worksheet=ws)
        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
        ws.column_dimensions = dim_holder

    def get_response(self, workbook):
        file_name = self.get_file_name()
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        workbook.save(response)
        return response
    

    def post(self, request, *args, **kwargs):
        self.before_process()
        workbook, ws = self.get_worksheet()
        self.before(ws)
        self.set_sheet_title(ws)
        self.set_header(ws)

        data, total_hours, days_checkeds = self.get_report_content() 

        for record in data:
            ws.append(self.process_row(record, days_checkeds=days_checkeds))

        ws.append(self.get_footer(total_hours))

        self.post_processing(data, ws, workbook)
        self.fit_column_size(ws)
        return self.get_response(workbook)
    

class ReportBrandMixin:

    def get_headlines(self, department, model_id):
        return ""
    
    def before(self, ws):
        now = timezone.now()
        ws.merge_cells("A1:A6")
        img = Image("/app/src/static/images/branding/logo_inpromaro_lit_backup.png")
        ws.add_image(img, "A1")
        ws.merge_cells("B1:G6")
        department = None
        if "department" in self.request.POST:
            department = Department.objects.get(id=self.request.POST.get("department"))

        else:
            department = Employee.objects.select_related("department").get(id=int(self.request.POST.get("employer"))).department
        
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B1"].font = Font(bold=True, size=13, name="Arial")

        ws["B1"].value = self.get_headlines(department, int(self.request.POST.get("department", self.request.POST.get("employer"))))
        for i in range(6):
            ws.append([" ", " ", " ", " ", " ", " ", " "])


class ReportWorkerExcel(ReportBrandMixin, ReportExcelMixin):

    def before_process(self):
        self.observations = DailyCalendarObservation.objects.select_related("employer", "calendar_day").filter(
            employer_id=int(self.request.POST.get("employer")),
            calendar_day__date_day__range=[ 
                self.request.POST.get("start_at"), 
                self.request.POST.get("end_at")
            ]
        )

    def get_headlines(self, department, model_id):
        now = timezone.now()
        
        employer = Employee.objects.get(id=model_id)
        start_at = datetime.strptime(self.request.POST.get("start_at"), "%Y-%m-%d")
        end_at = datetime.strptime(self.request.POST.get("end_at"), "%Y-%m-%d")
        return (
            " Reporte por departamento \n"
            f"Departamento: {department.name} \n"
            f"Trabajador: {employer.name} {employer.last_name} - {employer.cedula} \n"
            f"Desde {start_at.strftime('%d/%m/%Y')} hasta {end_at.strftime('%d/%m/%Y')} \n"
            "Fecha de generación: {0} \n ".format(now.strftime("%d/%m/%Y"))
        )
    
    def get_headers(self):
        return [
            "Fecha", 
            "Hora de entrada",
            "Hora de salida",
            "Observación",
            "Total de horas"
        ]
    
    def process_row(self, record, *args, **kwargs):
        observation = has_observation(record, self.observations)
        return [
            record["created"].strftime("%d/%m/%Y"),
            record["start_at"],
            record["end_at"],
            "Si" if observation else "No",
            record["abs_total_hours"]
        ]

    def get_sheet_title(self):
        return "Reporte por trabajador"

    def get_file_name(self):
        return "{at}_reporte_por_trabajador.xlsx".format(
            at=calendar.timegm(timezone.now().timetuple())
        )

    def get_report_content(self):
        return DailyChecks.objects.report_by_employee( int(self.request.POST.get("employer")), self.request.POST.get("start_at"), self.request.POST.get("end_at"))
    
    def post_processing(self, data, sheet, wb):
        observation_sheet = wb.create_sheet("Observaciones")
        observation_sheet.append([
            "Nombre y apellido",
            "Cédula",
            "Fecha",
            "Descripción",
        ])
        for observation in self.observations:
            observation_sheet.append([
                observation.employer.get_fullname(),
                observation.employer.cedula,
                observation.calendar_day.date_day.strftime("%d/%m/%Y"),
                observation.description
            ])

        dim_holder = DimensionHolder(worksheet=observation_sheet)
        for col in range(observation_sheet.min_column, observation_sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(observation_sheet, min=col, max=col, width=20)
        
    
class ReportDepartmentExcel(ReportBrandMixin, ReportExcelMixin):

    def before_process(self):
        self.observations = DailyCalendarObservation.objects.select_related("employer", "calendar_day").filter(
            employer__department_id=int(self.request.POST.get("department")),
            calendar_day__date_day__range=[ 
                self.request.POST.get("start_at"), 
                self.request.POST.get("end_at")
            ]
        )

    def get_headlines(self, department, model_id):
        now = timezone.now()
        start_at = datetime.strptime(self.request.POST.get("start_at"), "%Y-%m-%d")
        end_at = datetime.strptime(self.request.POST.get("end_at"), "%Y-%m-%d")
        return (
            " Reporte por departamento \n"
            f"Departamento: {department.name} \n"
            f"Desde {start_at.strftime('%d/%m/%Y')} hasta {end_at.strftime('%d/%m/%Y')} \n"
            "Fecha de generación: {0} \n ".format(now.strftime("%d/%m/%Y"))
        )
    
    def get_headers(self):
        return [
            "Cédula", 
            "Nombre",
            "Apellido",
            "Cargo",
            "Días Marcados",
            "Total de horas"
        ]
    
    def process_row(self, record, *args, **kwargs):
        observation = has_observation(record, self.observations)
        days_checkeds = kwargs.get("days_checkeds", None)[record["employer"].id] if kwargs.get("days_checkeds", None) is not None else "--"
        return [
            record["employer"].cedula,
            record["employer"].name,
            record["employer"].last_name,
            record["employer"].position.position,
            days_checkeds,
            record["abs_total_hours"]
        ]

    def get_sheet_title(self):
        return "Reporte por departamento"

    def get_file_name(self):
        return "{at}_reporte_por_departamento.xlsx".format(
            at=calendar.timegm(timezone.now().timetuple())
        )
    
    def get_report_content(self):
        return DailyChecks.objects.report_by_department(
            from_date=self.request.POST.get("start_at"), 
            until_date=self.request.POST.get("end_at"), 
            department=int(self.request.POST.get("department"))
        )
    
    
    def post_processing(self, data, sheet, wb):
        observation_sheet = wb.create_sheet("Observaciones")
        observation_sheet.append([
            "Nombre y apellido",
            "Cédula",
            "Fecha",
            "Descripción",
        ])
        for observation in self.observations:
            observation_sheet.append([
                observation.employer.get_fullname(),
                observation.employer.cedula,
                observation.calendar_day.date_day.strftime("%d/%m/%Y"),
                observation.description
            ])

        dim_holder = DimensionHolder(worksheet=observation_sheet)
        for col in range(observation_sheet.min_column, observation_sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(observation_sheet, min=col, max=col, width=20)
        
    


class ReportAttendanceExcel(ReportBrandMixin, ReportExcelMixin):

 
    def default(self, request, *args, **kwargs):
        self.before_process()
        workbook = Workbook()
        headers = self.get_headers()
        ws = workbook.active
        self.before(ws)
        ws.title = self.get_sheet_title()

        ws.append(headers)
        data = self.get_report_content() 

        for record in data:
            ws.append(self.process_row(record))

        self.post_processing(data, ws, workbook)


        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
        
        ws.column_dimensions = dim_holder
        file_name = self.get_file_name()
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        workbook.save(response)
        return response
    
    def before(self, ws):
        now = timezone.now()
        ws.merge_cells("A1:A6")
        img = Image("/app/src/static/images/branding/logo_inpromaro_lit_backup.png")
        ws.add_image(img, "A1")
        if self.request.GET.get("type_report", "simple") =="simple":
            ws.merge_cells("B1:G6")
        else:
            ws.merge_cells("B1:J6")
        department = None
        if int(self.request.GET.get("by_office", 0)) == 1:
            department = Department.objects.get(id=self.request.GET.get("department"))            
            ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["B1"].font = Font(bold=True, size=10, name="Arial")

        ws["B1"].value = self.get_headlines(department, int(self.request.GET.get("department", 0)))
        for i in range(6):
            ws.append([" ", " ", " ", " ", " ", " ", " "])
            
    def get_headlines(self, department, model_id):
        now = timezone.now()
        
        start_at = datetime.strptime(self.request.GET.get("start_at"), "%Y-%m-%d")
        end_at = datetime.strptime(self.request.GET.get("end_at"), "%Y-%m-%d")

        department_header = f"Departamento: {department.name} \n" if department is not None else ""
        return (
            " Reporte de asistencias \n"
            f"{department_header}"
            f"Desde {start_at.strftime('%d/%m/%Y')} hasta {end_at.strftime('%d/%m/%Y')} \n"
            "Fecha de generación: {0} \n ".format(now.strftime("%d/%m/%Y"))
        )
    
    def get_headers(self):
        headers = [
            "Cédula", 
            "Nombre",
            "Apellido",
            "Cargo",
        ]

        if self.request.GET.get("type_report", "simple") != "simple":
            headers.append("Fecha de entrada")
            headers.append("Hora de entrada")
            headers.append("Fecha de salida")
            headers.append("Hora de salida")
            headers.append("Horas trabajadas")
        else:
            headers.insert(0, "")
        
        headers.append("Departamento")
        return headers
    
    def process_row(self, record, *args, **kwargs):
        columns = []

        if self.request.GET.get("type_report", "simple") == "simple":
            columns =  [
                record['row'],
                record['employer']['cedula'],
                record['employer']['name'],
                record['employer']['last_name'],
                record['employer']['position']['position'],
            ]
        
        else:
            columns = [
                record['cedula'],
                record['name'],
                record['last_name'],
                record['position']['position'],
                "--",
                "--",
                "--",
                "--",
                "--",
            ]
        
        columns.append(record["employer"]['department'] if 'employer' in record else record['department'])
        return columns
    

    def get_sheet_title(self):
        return "Reporte por trabajador"

    def get_file_name(self):
        return "{at}_reporte_de_asistencia.xlsx".format(
            at=calendar.timegm(timezone.now().timetuple())
        )

    def get_records(self):
        from src.clocking.reportering import AttendanceReport
        department = None
        if int(self.request.GET.get("by_office", 0)) == 1:
            department = int(self.request.GET.get("department"))

        report = AttendanceReport(
            self.request.GET.get("start_at"), 
            self.request.GET.get("end_at"), 
            department=department, 
            include_unattendances=False,
            work_center=int(self.request.GET.get("work_center", 1))
        )

        if self.request.GET.get("type_report", "simple") == "simple":
            print("Records simple")
            return report.sort().unique_by_user().make(simple=True)[0]
        
        return dict(report.sort().group_by_user())
    
    def get(self, request, *args, **kwargs):
        weekdays = {
            0: "Lunes",
            1: "Martes",
            2: "Miércoles",
            3: "Jueves",
            4: "Viernes",
            5: "Sábado",
            6: "Domingo",
        }

        months = {
            1: "Enero",
            2: "Febrero",
            3: "Marzo",
            4: "Abril",
            5: "Mayo",
            6: "Junio",
            7: "Julio",
            8: "Agosto",
            9: "Septiembre",
            10: "Octubre",
            11: "Noviembre",
            12: "Diciembre",
        }
        if self.request.GET.get("type_report", "simple") == "simple":
            return self.default(request, *args, **kwargs)
        
        new_page = False
        workbook, ws = self.get_worksheet()
        first_page = ws
        self.before(ws)
        self.set_sheet_title(ws, self.get_sheet_title())
        self.set_header(ws)

        data = self.get_records()
        if self.request.GET.get("type_report", "simple") == "simple":
            for records in data:
                if new_page:
                    title = f"Detalle del día {records.strftime('%d-%m-%Y')}"
                    ws = workbook.create_sheet(title)
                    self.before(ws)
                    self.set_header(ws)
                    self.post_processing(data[records], ws, workbook)
                    self.fit_column_size(ws)

                for record in data[records]:
                    ws.append(self.process_row(record))
                new_page = True
        else:
            for record in data.values():
                current_record = self.process_row(record)
                checking_record = []
                ws.append(self.process_row(record))
                for time_record in record["dates"]:
                    checking_record = [ "", "", "", "",
                                    f"{weekdays[time_record.day.weekday()]} {time_record.day.strftime('%d')} {months[time_record.day.month]} de {time_record.day.year}",
                                    time_record.start_time.strftime("%I:%M") if time_record.start_time is not None else "",
                                    f"{weekdays[time_record.end_time.weekday()]} {time_record.end_time.strftime('%d')} {months[time_record.end_time.month]} de {time_record.end_time.year}" if time_record.end_time is not None else "",
                                    time_record.end_time.strftime("%I:%M") if time_record.end_time is not None else "",
                                    time_record.total_hours,
                                    ]
                    ws.append(checking_record)


        self.fit_column_size(first_page)
        return self.get_response(workbook)
        

    def get_report_content(self):
        self.context = {}
        records = self.get_records()

        if int(self.request.GET.get("by_office", 0)) == 1:
            self.context["department"] = Department.objects.filter(id=int(self.request.GET.get("department"))).first()

        self.context["data"] = records
        self.context["range_start"] = datetime.strptime(self.request.GET.get("start_at"), "%Y-%m-%d").strftime("%d/%m/%Y")
        self.context["range_end"] = datetime.strptime(self.request.GET.get("end_at"), "%Y-%m-%d").strftime("%d/%m/%Y")
        self.context["only_department"] = int(self.request.GET.get("by_office", 0)) == 1
        return self.context["data"]
    