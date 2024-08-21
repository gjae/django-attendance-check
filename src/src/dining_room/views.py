from django.shortcuts import render
from datetime import datetime
import time
from django.http.response import JsonResponse
from django.forms.models import model_to_dict
from django.views.generic import TemplateView
from django.http.response import HttpResponse
from django.db.models import Window, Q, Count
from django.db.models.functions import RowNumber

from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font

from src.dining_room.models import DiningChecking , ConfDiningRoom
from src.employees.models import Employee
from src.clocking.models import DailyChecks
from src.dining_room.managers import EmployerNotPresentException
from unfold.views import UnfoldModelAdminViewMixin

# Create your views here.

class ReportAdminView(UnfoldModelAdminViewMixin, TemplateView):
    title = "Gestion de reportes de comedor"  # required: custom page header title
    permissions_required = (
        "dining_room.view_dining_room",
    ) # required: tuple of permissions
    template_name = "dining_room/report_admin_template.html"


def report_dining_today_excel(request, *args, **kwargs):
    workbook = Workbook()
    ws = workbook.active
    current_date = datetime.now()
    today_data = DiningChecking.objects.select_related("employer", "conf_dining_room", "employer__department").filter(created__date=current_date.date(), conf_dining_room__isnull=False , conf_dining_room__is_removed=False).annotate(
        row_number=Window(
            RowNumber(),
            order_by=["employer__last_name", "employer__name", "-created"]
        )
    ).order_by("row_number", "conf_dining_room__check_name")

    total_by_benefit = (
        ConfDiningRoom
        .objects
        .values("check_name")
        .annotate(
            total_by_benefit=Count("checkings__id", filter=Q(checkings__created__date=current_date.date()))
        )
        .filter(total_by_benefit__gt=0)
        .values("total_by_benefit", "check_name")
    )

    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    aligment = Alignment(horizontal="center", vertical="center")
    font_subtitle = Font(name="Arial", size=14)
    font = Font(
        name="Calibri",
        size=11,
        bold=True
    )
    data_row_from = 4

    ws.merge_cells("A1:G1")
    ws.merge_cells("A3:G3")
    ws["A1"].value = f"PERSONAL ASISTENTE AL {current_date.strftime('%d/%m/%Y')} INPROMARCA"
    ws["A3"].value = "Trabajador"
    ws["C3"].border = thin_border
    ws["B3"].border = thin_border
    ws["D3"].border = thin_border
    ws["E3"].border = thin_border
    ws["F3"].border = thin_border
    ws["G3"].border = thin_border
    ws["A3"].border = thin_border

    ws["A1"].font = font
    ws["A3"].alignment = aligment
    ws["A1"].alignment= aligment
    ws["A3"].font= font_subtitle

    for data in today_data:
        ws[f"A{data_row_from}"].value = data.row_number
        ws[f"B{data_row_from}"].value = data.employer.cedula
        ws[f"C{data_row_from}"].value = data.employer.get_fullname()
        ws[f"E{data_row_from}"].value = data.employer.department.name
        ws[f"F{data_row_from}"].value = data.conf_dining_room.check_name
        ws[f"G{data_row_from}"].value = data.created.strftime("%d/%m/%Y %I:%M %p")

        ws.merge_cells(f"C{data_row_from}:D{data_row_from}")
        ws[f"A{data_row_from}"].border = thin_border
        ws[f"B{data_row_from}"].border = thin_border
        ws[f"C{data_row_from}"].border = thin_border
        ws[f"D{data_row_from}"].border = thin_border
        ws[f"E{data_row_from}"].border = thin_border
        ws[f"F{data_row_from}"].border = thin_border
        ws[f"G{data_row_from}"].border = thin_border

        ws[f"B{data_row_from}"].alignment = aligment
        ws[f"A{data_row_from}"].alignment = aligment
        data_row_from = data_row_from + 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 20
    response = HttpResponse(content_type="application/ms-excel")
    content = "attachment; filename =Comedor_{0}.xlsx".format(
        int(time.mktime(current_date.timetuple()))
    )

    data_row_from += 4
    ws.merge_cells(f"D{data_row_from}:E{data_row_from}")
    ws[f"C{data_row_from}"].value = "Beneficio"
    ws[f"D{data_row_from}"].value = "Total/Día"

    
    data_row_from += 1
    for benefit in total_by_benefit:
        ws[f"C{data_row_from}"].value = benefit["check_name"]
        ws[f"D{data_row_from}"].value = benefit["total_by_benefit"]
        ws[f"C{data_row_from}"].border = thin_border
        ws[f"D{data_row_from}"].border = thin_border
        data_row_from += 1
        
    response["Content-Disposition"] = content
    workbook.save(response)
    return response

def index(request, *args, **kwargs):
    today_checks = DiningChecking.objects.today_checks()

    return render(
        request, 
        "dining_room/index.html", 
        {"today_checks": today_checks}
    )

def default_today_last_checks(request, *args, **kwargs):
    today_checks = DiningChecking.objects.today_checks().order_by("-created")[0:39]
    response = []
    for check in today_checks:
        response.append({
            "name": check.employer.name,
            "lastname": check.employer.last_name,
            "position": check.employer.position.position if check.employer.position is not None else '',
            "department": check.employer.department.name if check.employer.department is not None else '',
            "id": check.employer.id,
            "avatar": check.employer.picture.url if check.employer is not None and check.employer.picture.name else '',
            "is_birthday": check.employer.is_birthday,
            "check_turn": check.conf_dining_room.check_name,
            "check_at": check.created.strftime("%I:%M %p")
        })
        
    return JsonResponse({
        "error": False,
        "data": response
    })


def check_dining_employer(request, card_id, *args, **kwargs):
    emp = Employee.objects.select_related("position", "department").filter(cedula=card_id).first()
    print(emp, card_id)
    if emp is None:
        return JsonResponse({
            "error": True, 
            "can_check": False, 
            "checked": False,
            "employer": None
        })
    
    try:
        check = DiningChecking.objects.make_check_if_can(emp)
    except EmployerNotPresentException:
        return JsonResponse({
            "error": True, 
            "can_check": True, 
            "checked": False,
            "employer": None
        })
    
    if check is None:
        return JsonResponse({
            "error": True, 
            "can_check": True, 
            "checked": False,
            "employer": None
        })
    
    return JsonResponse({
        "error": False,
        "can_check": True,
        "checked": True,
        "employer": {
            "name": emp.name,
            "lastname": emp.last_name,
            "position": emp.position.position if emp.position is not None else '',
            "department": emp.department.name if emp.department is not None else '',
            "id": emp.id,
            "avatar": emp.picture.url if emp is not None and emp.picture.name else '',
            "is_birthday": emp.is_birthday,
            "check_turn": check.conf_dining_room.check_name,
            "check_at": check.created.strftime("%I:%M %p")
        }
    })